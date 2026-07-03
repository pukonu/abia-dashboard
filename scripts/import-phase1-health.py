#!/usr/bin/env python3
"""Import the Primary Healthcare Phase 1 field assessment into Supabase.

Mapping (faithful to the workbook):
  DOMAIN XX blocks          -> domains under the Primary Healthcare thematic area
                               (weight from Domain Summary; Nigeria benchmark and
                               WHO/SDG target kept in the domain description)
  Question rows (code,      -> indicators (weight from sheet; options + rationale
  question, options,           in the description; scored 0-100 against target 100)
  weight, rationale)
  PHC response columns      -> entity-level results for Q2 2026, plus one
                               state-level row per question (mean across PHCs)

Scoring: response letters map to an equal-interval rubric where the best
option scores 100 and the worst scores 0 (4 options: A=100, B=66.7, C=33.3,
D=0; 3 options: A=100, B=50, C=0). Domain scores then roll up in-app as the
question-weight-weighted mean of indicator scores.
"""

from __future__ import annotations

import json
import re
import sys
import urllib.error
import urllib.request
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = Path(
    "/Users/prologic/Google Drive/ATSAC/Project/Abia State Health Dashboard Project/"
    "Abia State Executive Dashboard - Phase 1.xlsx"
)

THEMATIC_AREA_ID = "b92bd03f-9f06-465b-bf8b-64cfb1ecedf7"  # Primary Healthcare
MDA_ID = "9507ec53-9cb8-457b-b97c-491ba2e0dc95"  # ASPHCDA

# Column headers in the sheet (cols G..Q) -> canonical entity names in the DB
PHC_COLUMNS = [
    ("OSISIOMA PRIMARY HEALTH CENTER, WARD:URATTA OSISIOMS NGWA.", "Osisioma PHC"),
    ("ABAYI ARIARIA PHC", "Abayi Ariaria PHC"),
    ("MBUTUOMA PHC", "Mbutoma PHC"),
    ("UMUOCHAM", "Umuocham PHC"),
    ("EKEARO", "Ekearo PHC"),
    ("OBIARO", "Obiaro PHC"),
    ("OSOKWA", "Osokwa PHC"),
    ("AMAPUIFE", "Amapuife PHC"),
    ("ABA NORTH CENTRAL", "Aba North Central PHC"),
    ("OSUSU 1", "Osusu 1 PHC"),
    ("ST EUGENE", "St Eugene PHC"),
]


def load_env() -> dict[str, str]:
    env: dict[str, str] = {}
    for line in (ROOT / ".env").read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            key, value = line.split("=", 1)
            env[key.strip()] = value.strip().strip('"')
    return env


class Supabase:
    def __init__(self, url: str, key: str):
        self.base = url.rstrip("/") + "/rest/v1"
        self.key = key

    def _call(self, method: str, path: str, body=None, prefer: str | None = None):
        headers = {
            "apikey": self.key,
            "Authorization": f"Bearer {self.key}",
            "Content-Type": "application/json",
        }
        if prefer:
            headers["Prefer"] = prefer
        req = urllib.request.Request(
            f"{self.base}/{path}",
            data=json.dumps(body).encode() if body is not None else None,
            headers=headers,
            method=method,
        )
        try:
            with urllib.request.urlopen(req) as resp:
                raw = resp.read().decode()
                return json.loads(raw) if raw else None
        except urllib.error.HTTPError as err:
            raise RuntimeError(f"{method} {path} -> {err.code}: {err.read().decode()}") from err

    def select(self, path: str):
        return self._call("GET", path)

    def insert(self, table: str, rows: list[dict]):
        return self._call("POST", table, rows, prefer="return=representation")

    def upsert(self, table: str, rows: list[dict], on_conflict: str):
        return self._call(
            "POST",
            f"{table}?on_conflict={on_conflict}",
            rows,
            prefer="return=representation,resolution=merge-duplicates",
        )


def count_options(options_text) -> int:
    n = len(re.findall(r"(?:^|\n)\s*[a-e]\.", str(options_text or ""), re.I))
    return n if n >= 2 else 4


def letter_to_score(letter, n_opts: int) -> float | None:
    if not letter:
        return None
    idx = ord(str(letter).strip().upper()[0]) - ord("A")
    if idx < 0 or idx >= n_opts:
        return None
    return round(100 * (n_opts - 1 - idx) / (n_opts - 1), 2)


def parse_workbook():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Assessment Questions"]
    ws_sum = wb["Domain Summary"]

    summary: dict[int, dict] = {}
    for row in range(8, 20):
        num = ws_sum.cell(row, 2).value
        if num is None:
            continue
        summary[int(num)] = {
            "nigeria_raw": ws_sum.cell(row, 5).value,
            "target_raw": ws_sum.cell(row, 6).value,
            "domain_weight": float(ws_sum.cell(row, 9).value),
        }

    domain_rows = [
        r
        for r in range(1, ws.max_row + 1)
        if isinstance(ws.cell(r, 2).value, str)
        and re.match(r"DOMAIN\s+\d+", ws.cell(r, 2).value.strip(), re.I)
    ]

    domains = []
    for dr in domain_rows:
        m = re.match(r"DOMAIN\s+(\d+)\s+(.+)", ws.cell(dr, 2).value.strip(), re.I)
        num, name = int(m.group(1)), m.group(2).strip()
        sm = summary[num]

        questions = []
        row = dr + 3
        while row <= ws.max_row:
            code = ws.cell(row, 2).value
            if isinstance(code, str) and re.match(r"\d+\.\d+", code.strip()):
                options = ws.cell(row, 4).value
                n_opts = count_options(options)
                responses = {}
                for offset, (_, canonical) in enumerate(PHC_COLUMNS):
                    letter = ws.cell(row, 7 + offset).value
                    score = letter_to_score(letter, n_opts)
                    if score is not None:
                        responses[canonical] = score
                questions.append(
                    {
                        "code": code.strip(),
                        "question": str(ws.cell(row, 3).value or "").strip(),
                        "options": options,
                        "weight": float(ws.cell(row, 5).value or 0),
                        "rationale": ws.cell(row, 6).value,
                        "responses": responses,
                    }
                )
            elif isinstance(code, str) and (
                "subtotal" in code.lower() or re.match(r"DOMAIN\s+\d+", code.strip(), re.I)
            ):
                break
            row += 1

        domains.append(
            {
                "num": num,
                "name": name,
                "nigeria_raw": sm["nigeria_raw"],
                "target_raw": sm["target_raw"],
                "weight": sm["domain_weight"],
                "questions": questions,
            }
        )
    return domains


def weighted_mean(items: list[tuple[float, float]]) -> float | None:
    total_w = sum(w for _, w in items)
    return round(sum(s * w for s, w in items) / total_w, 2) if total_w else None


def main():
    env = load_env()
    sb = Supabase(env["NEXT_PUBLIC_SUPABASE_URL"], env["SUPABASE_SERVICE_ROLE_KEY"])
    domains = parse_workbook()

    entities = sb.select(f"entities?select=id,name&mda_id=eq.{MDA_ID}")
    entity_ids = {e["name"]: e["id"] for e in entities}
    missing = [c for _, c in PHC_COLUMNS if c not in entity_ids]
    if missing:
        raise RuntimeError(f"PHC entities missing in DB: {missing}")

    period = sb.upsert(
        "time_periods",
        [{"frequency": "quarterly", "label": "Q2 2026", "start_date": "2026-04-01", "end_date": "2026-06-30"}],
        on_conflict="frequency,start_date",
    )[0]

    domain_payload = [
        {
            "thematic_area_id": THEMATIC_AREA_ID,
            "name": f"{d['num']:02d} — {d['name']}",
            "description": f"Nigeria: {d['nigeria_raw']} · WHO/SDG target: {d['target_raw']}",
            "weight": d["weight"],
        }
        for d in domains
    ]
    domain_rows = sb.upsert("domains", domain_payload, on_conflict="thematic_area_id,name")
    domain_id_by_name = {r["name"]: r["id"] for r in domain_rows}

    indicator_payload = []
    for d in domains:
        domain_id = domain_id_by_name[f"{d['num']:02d} — {d['name']}"]
        for q in d["questions"]:
            description = str(q["options"] or "")
            if q["rationale"]:
                description += f"\n\nRationale: {q['rationale']}"
            indicator_payload.append(
                {
                    "domain_id": domain_id,
                    "name": f"{q['code']} {q['question']}",
                    "description": description.strip(),
                    "unit": "%",
                    "direction": "higher_is_better",
                    "target_value": 100,
                    "target_source": "WHO / SDG framework",
                    "weight": q["weight"],
                }
            )
    indicator_rows = sb.upsert("indicators", indicator_payload, on_conflict="domain_id,name")
    indicator_id_by_name = {r["name"]: r["id"] for r in indicator_rows}

    result_payload = []
    for d in domains:
        for q in d["questions"]:
            indicator_id = indicator_id_by_name[f"{q['code']} {q['question']}"]
            scores = list(q["responses"].values())
            if scores:
                result_payload.append(
                    {
                        "indicator_id": indicator_id,
                        "time_period_id": period["id"],
                        "entity_id": None,
                        "abia_value": round(sum(scores) / len(scores), 2),
                        "target_value": 100,
                        "notes": f"Mean of {len(scores)} PHC field responses · Q2 2026 assessment",
                    }
                )
            for canonical, score in q["responses"].items():
                result_payload.append(
                    {
                        "indicator_id": indicator_id,
                        "time_period_id": period["id"],
                        "entity_id": entity_ids[canonical],
                        "abia_value": score,
                        "target_value": 100,
                        "notes": "Q2 2026 PHC field assessment",
                    }
                )

    for i in range(0, len(result_payload), 500):
        sb.insert("results", result_payload[i : i + 500])

    # ---- scoring report ----
    print(f"Imported: {len(domain_rows)} domains, {len(indicator_rows)} indicators, "
          f"{len(result_payload)} results into {period['label']}\n")

    print(f"{'Domain':<44}{'Weight':>7}{'Score':>8}  Nigeria / WHO-SDG target")
    print("-" * 100)
    overall = []
    for d in domains:
        items = []
        for q in d["questions"]:
            scores = list(q["responses"].values())
            if scores:
                items.append((sum(scores) / len(scores), q["weight"]))
        score = weighted_mean(items)
        overall.append((score, d["weight"]))
        print(
            f"{d['num']:02d} — {d['name']:<38}{d['weight']:>7.3f}{score:>8.1f}  "
            f"{d['nigeria_raw']} / {d['target_raw']}"
        )
    total = weighted_mean([(s, w) for s, w in overall if s is not None])
    print("-" * 100)
    print(f"{'PRIMARY HEALTHCARE — OVERALL (weighted)':<44}{1.0:>7.3f}{total:>8.1f}")

    print("\nPer-PHC overall scores (domain-weighted):")
    phc_totals = []
    for _, canonical in PHC_COLUMNS:
        domain_items = []
        for d in domains:
            q_items = [
                (q["responses"][canonical], q["weight"])
                for q in d["questions"]
                if canonical in q["responses"]
            ]
            ds = weighted_mean(q_items)
            if ds is not None:
                domain_items.append((ds, d["weight"]))
        t = weighted_mean(domain_items)
        if t is not None:
            phc_totals.append((canonical, t))
    for canonical, t in sorted(phc_totals, key=lambda x: -x[1]):
        print(f"  {canonical:<28}{t:>7.1f}")


if __name__ == "__main__":
    main()
